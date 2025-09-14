from copy import deepcopy
import scrapy
from scrapy import signals
from scrapy.utils.response import open_in_browser


class ArmytrixSpiderSpider(scrapy.Spider):
    name = "armytrix_spider"
    start_urls = ["https://www.armytrix.com/product-exhaust"]
    fields = ['Url', 'Brand', 'Model', 'Engin', 'Title', 'FITMENT', 'NOTE', 'FEATURE', 'SKU', 'Title1', 'Title1 Name',
              'Title2', 'Variant Image', 'Image1', 'Image2', 'Image3', 'Image4', 'Image5', 'Image6',
              'Image7', 'Image8', 'Image9', 'Image10',
              'Image11', 'Image12', 'Image13', 'Image14', 'Image15', 'Image16', 'Image17', 'Image18', 'Image19',
              'Image20', 'Image21', 'Image22', 'Image23', 'Image24', 'Image25', 'Image26', 'Image27', 'Image28',
              'Image29', 'Image30', 'Image31', 'Image32', 'Image33', 'Image34', 'Image35', 'Image36',
              'Image37', 'Image38', 'Image39', 'Image40', 'Image41', 'Image42', 'Image43', 'Image44', 'Image45']
    custom_settings = {
        'ROBOTSTXT_OBEY': False,
        'CONCURRENT_REQUESTS': 1,
        # 'DOWNLOAD_DELAY': 1,
        'FEEDS': {'output/army_data_new_2024.csv': {'format': 'csv', 'fields': fields}}
    }
    headers = {
        'authority': 'www.armytrix.com',
        'accept': '*/*',
        'accept-language': 'en-US,en;q=0.9',
        'cache-control': 'no-cache',
        'content-type': 'application/x-www-form-urlencoded; charset=UTF-8',
        # 'cookie': 'CakeCookie[guest_id]=1702970635; _fbp=fb.1.1702970642236.974389390; _gcl_au=1.1.1919071710.1702970642; _gid=GA1.2.1044417050.1702970643; CAKEPHP=b8647bb08bcb3c737a78fe511d0ac269; _ga=GA1.2.682853458.1702970641; _ga_DHWDMB6JQ1=GS1.1.1703067351.4.1.1703068433.34.0.0',
        'dnt': '1',
        'origin': 'https://www.armytrix.com',
        'pragma': 'no-cache',
        'referer': 'https://www.armytrix.com/product-exhaust',
        'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'x-requested-with': 'XMLHttpRequest',
    }
    data = {
        'type': 'car',
        'make_id': '',
        'target_id': 'RequestModel',
    }

    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super(ArmytrixSpiderSpider, cls).from_crawler(crawler, *args, **kwargs)
        crawler.signals.connect(spider.spider_closed_army, signal=signals.spider_closed)
        return spider

    final_list = []

    def start_requests(self):
        yield scrapy.Request(url=self.start_urls[0], headers=self.headers)

    def parse(self, response, **kwargs):
        for brand in response.css('select#RequestBrand>option:not(:first-child)'):
            brand_name = brand.css('::text').get('').strip()
            brand_value = brand.css('::attr(value)').get('')
            data = deepcopy(self.data)
            data['make_id'] = brand_value
            yield scrapy.FormRequest(url='https://www.armytrix.com/pages/get_for', headers=self.headers, formdata=data,
                                     callback=self.parse_model,
                                     meta={'brand_value': brand_value, 'brand_name': brand_name})

    def parse_model(self, response):
        headers = {
            'authority': 'www.armytrix.com',
            'accept': '*/*',
            'accept-language': 'en-US,en;q=0.9',
            'cache-control': 'no-cache',
            'content-type': 'application/x-www-form-urlencoded; charset=UTF-8',
            # 'cookie': 'CakeCookie[guest_id]=1702970635; _fbp=fb.1.1702970642236.974389390; _gcl_au=1.1.1919071710.1702970642; _gid=GA1.2.1044417050.1702970643; CAKEPHP=b8647bb08bcb3c737a78fe511d0ac269; _ga=GA1.2.682853458.1702970641; _ga_DHWDMB6JQ1=GS1.1.1703067351.4.1.1703068433.34.0.0',
            'dnt': '1',
            'origin': 'https://www.armytrix.com',
            'pragma': 'no-cache',
            'referer': 'https://www.armytrix.com/product-exhaust',
            'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'x-requested-with': 'XMLHttpRequest',
        }
        model_html = scrapy.Selector(text=response.text.split('html(')[-1].split(');</script')[0])
        for model in model_html.css("option:not(:first-child)"):
            model_name = model.css('::text').get('').strip()
            model_value = model.css('::attr(value)').get('')
            data = deepcopy(self.data)
            data['model_id'] = model_value
            yield scrapy.FormRequest(url='https://www.armytrix.com/pages/get_for', headers=headers, formdata=data,
                                     callback=self.parse_engin, meta={'model_value': model_value, 'brand_value':
                    response.meta['brand_value'], 'model_name': model_name, 'brand_name': response.meta['brand_name']})

    def parse_engin(self, response):
        engin_html = scrapy.Selector(text=response.text.split('html(')[-1].split(');</script')[0])
        for engin in engin_html.css("option:not(:first-child)"):
            engin_name = engin.css('::text').get('').strip()
            engin_value = engin.css('::attr(value)').get('')
            data = {
                'brand': f'{response.meta["brand_value"]}',
                'model': f'{response.meta["model_value"]}',
                'motor': f'{engin_value}'}
            yield scrapy.FormRequest(url='https://www.armytrix.com/pages/check_product', headers=self.headers,
                                     formdata=data, callback=self.detail_page_url,
                                     meta={'brand_name': response.meta['brand_name'],
                                           'model_name': response.meta['model_name'], 'engin_name': engin_name})

    def detail_page_url(self, response):
        page_url = response.text.split("='", 1)[-1].split("';", 1)[0].strip()
        yield scrapy.Request(url=response.urljoin(page_url), headers=self.headers, callback=self.detail_page,
                             meta=response.meta)
        # yield scrapy.Request(url='https://www.armytrix.com/product/toyota-supra-mk5-b58b30-electic-valve',
        #                      headers=self.headers, callback=self.detail_page, meta=response.meta)

    def detail_page(self, response):
        # open_in_browser(response)
        item = dict()
        item['Url'] = response.url
        item['Brand'] = response.meta['brand_name']
        item['Model'] = response.meta['model_name']
        item['Engin'] = response.meta['engin_name']
        item['Title'] = response.css('h1::text').get('').strip()
        item['FITMENT'] = ' '.join(
            response.xpath("//h2[contains(text(),'FITMENT')]/following-sibling::text()").getall()).strip()
        item['NOTE'] = ' '.join(
            response.xpath("//h2[contains(text(),'NOTE')]/following-sibling::text()").getall()).strip()
        item['FEATURE'] = ' '.join(
            response.xpath("//h2[contains(text(),'FEATURE')]/following-sibling::text()").getall()).strip()
        images = response.css('div#product_slider>div.prodctBg>img::attr(src)').getall()
        for image in images:
            item[f"Image{images.index(image) + 1}"] = f'https:{image.strip()}'
        ids_of_dropdows = ['cat_back_ul', 'tuning_ul', 'ecu_ul']
        for dropdown_id in ids_of_dropdows:
            for dropdown in response.xpath(f"//ul[@id='{dropdown_id}']/li[position()>1]"):
                item['SKU'] = dropdown.xpath(".//div/h4[not(@class)]/text()").get('').strip()
                item['Title1'] = ' '.join(
                    dropdown.xpath(".//div[@class='clearfix']/following-sibling::text()").getall()).strip()
                item['Title1 Name'] = dropdown.xpath("./ancestor::div[contains(@class,'add-bx-sel')]/h2/text()").get(
                    '').strip()
                item['Title2'] = dropdown.xpath(".//div/h4[@class]/text()").get('').strip()
                item['Variant Image'] = dropdown.xpath("./@full_img").get('').strip()
                self.final_list.append(item.copy())
                yield item

    def spider_closed_army(self, spider):
        self.write_excel_file1(self.final_list)

    def write_excel_file1(self, items):
        import openpyxl
        filename = 'army_data_new_2024'
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        fieldnames = ['Url', 'Brand', 'Model', 'Engin', 'Title', 'FITMENT', 'NOTE', 'FEATURE', 'SKU', 'Title1',
                      'Title1 Name',
                      'Title2', 'Variant Image', 'Image1', 'Image2', 'Image3', 'Image4', 'Image5', 'Image6',
                      'Image7', 'Image8', 'Image9', 'Image10',
                      'Image11', 'Image12', 'Image13', 'Image14', 'Image15', 'Image16', 'Image17', 'Image18', 'Image19',
                      'Image20', 'Image21', 'Image22', 'Image23', 'Image24', 'Image25', 'Image26', 'Image27', 'Image28',
                      'Image29', 'Image30', 'Image31', 'Image32', 'Image33', 'Image34', 'Image35', 'Image36',
                      'Image37', 'Image38', 'Image39', 'Image40', 'Image41', 'Image42', 'Image43', 'Image44', 'Image45']
        for i in items:
            fields = i.keys()
            for field in fields:
                if field not in fieldnames:
                    fieldnames.append(field)
        for col_idx, fieldname in enumerate(fieldnames, start=1):
            sheet.cell(row=1, column=col_idx, value=fieldname)
        for row_idx, record in enumerate(items, start=2):
            for col_idx, fieldname in enumerate(fieldnames, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=record.get(fieldname, ''))
        workbook.save(f"{filename}.xlsx")
